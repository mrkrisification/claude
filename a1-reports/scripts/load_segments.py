#!/usr/bin/env python3
"""
Load per-country (segment) financials into data/financials.duckdb.

Source: A1 Group "Analyst Fact Sheets" (Excel), downloaded under factsheets/.
These break the group into its reportable country segments — Austria, Bulgaria,
Croatia, Belarus, Slovenia, Serbia, North Macedonia — plus Corporate/Eliminations.
The annual reports only ever publish the bundled "Additional Markets" aggregate;
the factsheets are the only source that splits Slovenia / Serbia / Macedonia out.

What we load (per fiscal year, per segment):
  - revenue  : segment "Revenues" (total revenues incl. other operating income)
  - ebitda   : segment "EBITDA comparable" for 2010-2015 (matches the group
               `ebitda` basis those years), plain "EBITDA" from 2016 on.
Both reconcile to the group `segment='total'` rows already in the DB (Σ countries
+ corporate ≈ group, within rounding/eliminations).

Provenance: factsheet rows have no PDF page, so source_page is NULL and `notes`
records the factsheet file + sheet. source_year = fiscal_year + 1 (publication),
matching the loader convention in load_financials.py. restated_flag = FALSE
(each year is taken as first reported in its own year-end factsheet).

Layout families handled:
  - 2010-2021 : sheet "Results by Segments"
  - 2022-2025 : sheet "Operating Results by Segment_h"
(2007-2009 use a mobile-only "Mobile Communication" sheet — not loaded here;
 see notes in wiki. 2019's Q4 factsheet was a PDF, so FY2019 is read from the
 FS_FY2020 file's FY2019 column.)
"""
import os, re, glob
import duckdb
import openpyxl
import xlrd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
FS   = os.path.join(ROOT, "factsheets")
DB   = os.path.join(ROOT, "data", "financials.duckdb")

CTYKEYS = [("Austria","austria"),("Bulgaria","bulgaria"),("Croatia","croatia"),
           ("Belarus","belarus"),("Slovenia","slovenia"),("Serbia","serbia"),
           ("Macedonia","north_macedonia")]
EXCL = ("group","total","elimination","consolidat")
SHEETS = ["Operating Results by Segment_h", "Results by Segments"]

# fiscal_year -> (factsheet file, publication/source_year)
def fs_path(year):
    hits = glob.glob(os.path.join(FS, f"FS_FY{year}.xls*"))
    return hits[0] if hits else None

# year -> (file to read, source_year). 2019 comes from the FY2020 factsheet.
def plan():
    p = {y: (fs_path(y), y+1) for y in range(2010, 2026) if fs_path(y) and y != 2019}
    if fs_path(2020):
        p[2019] = (fs_path(2020), 2020)   # FY2019 column lives in the FS_FY2020 file
    return p

def rows_of(path, sheet):
    if path.endswith(".xls"):
        wb = xlrd.open_workbook(path); s = wb.sheet_by_name(sheet)
        return [s.row_values(r) for r in range(s.nrows)]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return list(wb[sheet].iter_rows(values_only=True))

def pick_sheet(path):
    names = ([s.name for s in xlrd.open_workbook(path).sheets()] if path.endswith(".xls")
             else openpyxl.load_workbook(path, read_only=True).sheetnames)
    for n in SHEETS:
        if n in names:
            return n
    raise ValueError(f"no segment sheet in {path}")

def fy_cells(strs):
    return {int(re.search(r"\d{4}", s).group()): j
            for j, s in enumerate(strs) if re.fullmatch(r"FY ?\d{4}", s)}

def label_of(strs, fyc):
    for j, s in enumerate(strs):
        if j in fyc:
            continue
        if s and not re.match(r"(Q\d|1-|H\d|in EUR|EUR|%|\d)", s):
            return s
    return ""

def metric_of(label):
    l = label.lower().strip()
    if label.strip() == "Revenues":
        return "revenue"
    if l.startswith("ebitda") and not any(x in l for x in
            ("after","margin","before","incl","restructuring","impairment")):
        return "ebitda"   # "EBITDA comparable" (2010-15) or plain "EBITDA" (2016+)
    return None

def seg_of(label):
    l = label.lower()
    if "corporate" in l or "elimination" in l:
        return "corporate"
    if any(x in l for x in EXCL):
        return None
    for kw, code in CTYKEYS:
        if kw.lower() in l:
            return code
    return None

def parse(path, want_year):
    sheet = pick_sheet(path)
    rows = rows_of(path, sheet)
    out, cur, fyc = {}, None, {}
    for r in rows:
        strs = [("" if c is None else str(c).strip()) for c in r]
        fc = fy_cells(strs)
        if len(fc) >= 2:                       # block header row
            cur = metric_of(label_of(strs, fc)); fyc = fc; continue
        if cur and want_year in fyc:
            seg = seg_of(label_of(strs, fyc))
            if seg:
                v = r[fyc[want_year]]
                if isinstance(v, (int, float)):
                    out[(cur, seg)] = round(float(v), 1)
    return out, sheet

def main():
    con = duckdb.connect(DB)
    # Group totals already in the DB, for the reconciliation print-out.
    gt = {(fy, m): v for fy, m, v in con.execute(
        "SELECT fiscal_year, metric_name, value FROM financials "
        "WHERE segment='total' AND restated_flag=FALSE "
        "AND metric_name IN ('revenue','ebitda')").fetchall()}

    n = 0
    for year, (path, src_year) in sorted(plan().items()):
        parsed, sheet = parse(path, year)
        note = f"A1 analyst factsheet {os.path.basename(path)} [{sheet}]"
        for (metric, seg), value in parsed.items():
            con.execute(
                "DELETE FROM financials WHERE fiscal_year=? AND metric_name=? "
                "AND segment=? AND source_year=? AND restated_flag=FALSE",
                [year, metric, seg, src_year])
            con.execute(
                "INSERT INTO financials "
                "(fiscal_year, metric_name, segment, value, unit, source_year, "
                " source_page, restated_flag, notes) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                [year, metric, seg, value, "EUR_million", src_year, None, False, note])
            n += 1
        rev = sum(v for (m, s), v in parsed.items() if m == "revenue")
        eb  = sum(v for (m, s), v in parsed.items() if m == "ebitda")
        print(f"  FY{year}: loaded {len(parsed):2} rows | Σrev={rev:7.1f} "
              f"(grp {gt.get((year,'revenue'))}) | Σebitda={eb:7.1f} "
              f"(grp {gt.get((year,'ebitda'))})")
    print(f"\nInserted/updated {n} segment rows.")
    tot = con.execute("SELECT count(*) FROM financials WHERE segment!='total'").fetchone()[0]
    print(f"financials now holds {tot} non-total (segment) rows.")
    con.close()

if __name__ == "__main__":
    main()
